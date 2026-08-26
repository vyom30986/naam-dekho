"""Build the connector catalog Excel workbook for the dev team."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Naam_Dekho_Connector_Catalog.xlsx")

# ── Palette ─────────────────────────────────────────────────────
INK = "0F1419"; INK2 = "3D4751"; INK3 = "6B7480"
ACCENT = "B8501C"; GOLD = "E8C76A"; BG = "FAF8F3"; BG2 = "F3EFE5"
OK_BG = "E7F2E9"; OK_INK = "1B5E20"
NO_BG = "FCE4EC"; NO_INK = "880E4F"
WARN_BG = "FFF4D9"; WARN_INK = "8A5A00"

# ── Styles ──────────────────────────────────────────────────────
thin = Side(border_style="thin", color="D8D0BC")
border = Border(top=thin, bottom=thin, left=thin, right=thin)

def style_header(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color="FAF8F3")
    cell.fill = PatternFill("solid", start_color=INK)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = border

def style_category(cell):
    cell.font = Font(name="Arial", size=11, bold=True, color="FAF8F3")
    cell.fill = PatternFill("solid", start_color=ACCENT)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = border

def style_body(cell, fill=None):
    cell.font = Font(name="Arial", size=10, color=INK)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = border
    if fill:
        cell.fill = PatternFill("solid", start_color=fill)


# ── Workbook ────────────────────────────────────────────────────
wb = Workbook()

# ════════════════════════════════════════════════════════════════
# SHEET 1 — CONNECTOR CATALOG
# ════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Connector Catalog"

# Title row
ws["A1"] = "Naam Dekho — Connector Catalog"
ws["A1"].font = Font(name="Arial", size=16, bold=True, color=INK)
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells("A1:I1")
ws.row_dimensions[1].height = 28

ws["A2"] = "Single source of truth for every external connector used by the platform. Update this file when a new source is added or pricing changes."
ws["A2"].font = Font(name="Arial", size=10, italic=True, color=INK3)
ws.merge_cells("A2:I2")
ws.row_dimensions[2].height = 18

# Header row
headers = [
    "Category", "Source / Connector", "Primary URL",
    "Access Method", "Free Tier", "Paid Tier (₹/USD)",
    "Rate Limit", "Auth Type", "Implementation Notes"
]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=h)
    style_header(cell)
ws.row_dimensions[4].height = 32

# Data
rows = [
    # ──────────────── LEGAL & REGULATORY ────────────────
    ("CATEGORY", "LEGAL & REGULATORY  —  Government registers (India)", "", "", "", "", "", "", ""),
    ("Legal", "MCA21 — Ministry of Corporate Affairs",
     "https://www.mca.gov.in/mcafoportal/showCheckCompanyName.do",
     "Web scraping (no public API)", "Free (manual)", "—",
     "~3 req/sec advisable", "Session cookie + CAPTCHA",
     "Use Playwright with warmed sessions. CAPTCHA via 2Captcha for deep tier."),
    ("Legal", "IP India — Trademark register (45 classes)",
     "https://ipindiaonline.gov.in/tmrpublicsearch/frmmain.aspx",
     "Web scraping (no public API)", "Free (manual)", "—",
     "~2 req/sec advisable", "Session + CAPTCHA",
     "Wordmark + Phonetic search. CAPTCHA solving via Anti-Captcha for deep tier."),
    ("Legal", "Copyright Office — Register of Copyrights",
     "https://copyright.gov.in/CRPublicSearch/PublicSearch.aspx",
     "Web scraping", "Free", "—",
     "~2 req/sec advisable", "None / session",
     "Search by title. Slower endpoint — cache results 7 days."),
    ("Legal", "GST common portal (trade-name search)",
     "https://services.gst.gov.in/services/searchtp",
     "Web scraping", "Free", "—",
     "~5 req/sec advisable", "Session cookie",
     "Per-state search. Use state code matrix. Cache 24h."),
    ("Legal", "DPIIT Startup India recognition register",
     "https://www.startupindia.gov.in/content/sih/en/search.html",
     "Web scraping", "Free", "—",
     "~5 req/sec advisable", "None",
     "JSON endpoint available via XHR — reverse-engineer the AJAX call."),
    ("Legal", "FSSAI brand-name database",
     "https://foscos.fssai.gov.in/",
     "Web scraping", "Free", "—",
     "~3 req/sec advisable", "Session",
     "Relevant only for food/beverage clients. Optional category gate."),
    ("Legal", "RBI — Banks / NBFCs master list",
     "https://www.rbi.org.in/Scripts/BankSearchMaster.aspx",
     "Web scraping + static CSV", "Free", "—",
     "Static — refresh daily", "None",
     "Download CSV nightly. Search in-memory."),
    ("Legal", "SEBI registered intermediaries",
     "https://www.sebi.gov.in/sebiweb/other/OtherAction.do?doRecognised=yes",
     "Web scraping + static CSV", "Free", "—",
     "Static — refresh weekly", "None",
     "Bulk export available. Index by company name."),
    ("Legal", "IRDAI insurer register",
     "https://www.irdai.gov.in/",
     "Web scraping + static CSV", "Free", "—",
     "Static — refresh monthly", "None",
     "Tiny dataset. In-memory fuzzy match."),
    ("Legal", "Patent Office — Applicant index",
     "https://search.ipindia.gov.in/IPOPatentSearch/PatentSearch/",
     "Web scraping", "Free", "—",
     "~2 req/sec", "Session + CAPTCHA",
     "Useful for prior-use evidence. CAPTCHA-gated."),
    ("Legal", "Trademarkia API (3rd-party fallback)",
     "https://www.trademarkia.com/api",
     "REST API", "—", "USD 99/month basic",
     "1000 req/day", "API key",
     "Useful when IP India is down. Coverage incomplete for Indian classes."),

    # ──────────────── DOMAINS ────────────────
    ("CATEGORY", "DOMAINS  —  Availability & pricing", "", "", "", "", "", "", ""),
    ("Domains", "WHOIS (raw protocol)",
     "https://www.iana.org/whois",
     "TCP port 43", "Free", "—",
     "~5 req/sec per server", "None",
     "Standard whois server map. Use python-whois or node-whois library."),
    ("Domains", "RDAP (modern WHOIS replacement)",
     "https://rdap.org/",
     "REST/JSON", "Free", "—",
     "Generous (varies)", "None",
     "Preferred over WHOIS where available. JSON response is structured."),
    ("Domains", "GoDaddy API",
     "https://developer.godaddy.com/",
     "REST API", "Free dev tier", "Production requires affiliate agreement",
     "60 req/min", "API key + secret",
     "Use for live price + availability. Affiliate program for revenue share."),
    ("Domains", "Namecheap API",
     "https://www.namecheap.com/support/api/",
     "REST API", "Free (requires $50 account balance OR 50 domains/year)", "—",
     "30 req/min", "API key + whitelisted IP",
     "Cheaper TLDs. Whitelist all production IPs."),
    ("Domains", "WhoisXML API",
     "https://www.whoisxmlapi.com/",
     "REST API", "500 req/month free", "USD 19+/month from 1k req",
     "Tier-dependent", "API key",
     "Use as fallback. Good for bulk WHOIS lookups."),
    ("Domains", "Domainr API",
     "https://domainr.com/docs/api",
     "REST API", "—", "USD 25/month from 10k req",
     "Tier-dependent", "API key (RapidAPI)",
     "Best UX for availability + suggestions. Pay-as-you-go via RapidAPI."),
    ("Domains", "INRegistry (.in / .co.in)",
     "https://registry.in/",
     "Web (no public API)", "Free WHOIS", "—",
     "Use WHOIS protocol", "None",
     "Use whois.registry.in via port 43."),
    ("Domains", "Cloudflare Registrar (at-cost)",
     "https://api.cloudflare.com/",
     "REST API", "Free tier", "At-cost domain pricing",
     "1200 req/5min", "API token",
     "Cheap registration if user has CF account. Limited TLDs."),

    # ──────────────── SOCIAL HANDLES ────────────────
    ("CATEGORY", "SOCIAL HANDLES  —  Username availability", "", "", "", "", "", "", ""),
    ("Social", "Instagram Graph API",
     "https://developers.facebook.com/docs/instagram-api/",
     "REST API", "Free with Meta dev account", "—",
     "200 req/hr (basic)", "OAuth 2.0",
     "For handle resolution use public profile URL HTTP HEAD check, not Graph API."),
    ("Social", "X (Twitter) API v2",
     "https://developer.x.com/",
     "REST API", "Limited (1500 reads/month Free tier)", "USD 200/month Basic; USD 5000+ Pro",
     "Tier-dependent", "OAuth 2.0 / Bearer",
     "Use lookup by username endpoint. Free tier insufficient for production volume."),
    ("Social", "YouTube Data API v3",
     "https://developers.google.com/youtube/v3",
     "REST API", "10,000 units/day free", "—",
     "1 unit per channel lookup", "API key",
     "Channel handle search. Cache aggressively."),
    ("Social", "LinkedIn API",
     "https://developer.linkedin.com/",
     "REST API", "Restricted partner access only", "Partner tier",
     "Partner-tier", "OAuth 2.0",
     "No public company-name availability API. Use unauthenticated URL probe."),
    ("Social", "Facebook Graph API",
     "https://developers.facebook.com/docs/graph-api/",
     "REST API", "Free with Meta dev account", "—",
     "200 req/hr", "OAuth 2.0",
     "Page slug availability via HTTP probe on facebook.com/<slug>."),
    ("Social", "Threads API (Meta)",
     "https://developers.facebook.com/docs/threads/",
     "REST API", "Free (preview)", "—",
     "Per Meta dev limits", "OAuth 2.0",
     "Limited preview. Fall back to public URL probe."),
    ("Social", "Telegram Bot API",
     "https://core.telegram.org/bots/api",
     "REST API", "Free", "—",
     "30 req/sec global", "Bot token",
     "getChat method returns channel/group existence. Free."),
    ("Social", "WhatsApp Business API (Meta)",
     "https://business.whatsapp.com/products/business-platform",
     "REST API", "Free dev sandbox", "USD 0.005–0.05 per conversation",
     "Per Meta limits", "OAuth + business verification",
     "Display-name availability via business directory."),
    ("Social", "Pinterest API",
     "https://developers.pinterest.com/",
     "REST API", "Free", "—",
     "200 req/hr standard", "OAuth 2.0",
     "Business profile slug check."),

    # ──────────────── MARKETPLACES ────────────────
    ("CATEGORY", "MARKETPLACES & STORES  —  App stores, e-commerce", "", "", "", "", "", "", ""),
    ("Marketplace", "google-play-scraper (npm)",
     "https://github.com/facundoolano/google-play-scraper",
     "Web scraping library", "Free / open source", "—",
     "Self-throttle to ~1 req/sec", "None",
     "Reliable. Use search() then exact-match check on title field."),
    ("Marketplace", "Apple iTunes Search API",
     "https://itunes.apple.com/search?term=&entity=software",
     "REST/JSON (public)", "Free", "—",
     "20 req/min", "None",
     "Public endpoint, no key. Filter by trackName."),
    ("Marketplace", "Product Hunt API v2 (GraphQL)",
     "https://api.producthunt.com/v2/docs",
     "GraphQL", "Free with OAuth", "—",
     "450 req/15min", "OAuth 2.0",
     "Strict on rate. Cache product names 1 day."),
    ("Marketplace", "GitHub REST API",
     "https://docs.github.com/en/rest",
     "REST/JSON", "5000 req/hr (auth)", "Enterprise plans",
     "5000 req/hr authenticated", "Personal access token",
     "GET /users/<name> returns 404 if free, 200 if taken."),
    ("Marketplace", "Shopify subdomain probe",
     "https://shopify.dev/api",
     "DNS / HTTP HEAD", "Free", "—",
     "Self-throttle", "None",
     "Resolve <name>.myshopify.com — 200 = taken, NXDOMAIN = free."),
    ("Marketplace", "Amazon (3rd-party scraping)",
     "https://www.amazon.in/",
     "Web scraping via ScrapingBee or Bright Data",
     "—", "USD 49+/month (ScrapingBee)",
     "Provider-dependent", "Provider API key",
     "No public seller-search API. Use ScrapingBee with rendered JS."),
    ("Marketplace", "Flipkart (3rd-party scraping)",
     "https://www.flipkart.com/",
     "Web scraping via Bright Data",
     "—", "USD 500+/month enterprise",
     "Provider-dependent", "Provider API key",
     "Brand Registry lookup. Heavy bot detection — needs residential proxies."),

    # ──────────────── BRAND COLLISION & SEO ────────────────
    ("CATEGORY", "BRAND COLLISION & SEO  —  SERP, news, knowledge", "", "", "", "", "", "", ""),
    ("Brand/SEO", "Google Custom Search JSON API",
     "https://developers.google.com/custom-search/v1/overview",
     "REST/JSON", "100 queries/day free", "USD 5 per 1k queries (up to 10k/day)",
     "10k req/day max", "API key + CX ID",
     "Use for exact-match SERP probe. Filter by gl=in (India locale)."),
    ("Brand/SEO", "SerpAPI",
     "https://serpapi.com/",
     "REST/JSON", "100 searches/month free", "USD 50/month from 5k",
     "Tier-dependent", "API key",
     "Easier than Google CSE for full SERP. Has Google Trends endpoint too."),
    ("Brand/SEO", "ScrapingBee",
     "https://www.scrapingbee.com/",
     "Proxy-rotating scraper", "1000 free credits", "USD 49+/month",
     "Per-credit", "API key",
     "Use for headless-render-needed sources (Amazon, dynamic pages)."),
    ("Brand/SEO", "Bright Data",
     "https://brightdata.com/",
     "Residential proxies + Web Unlocker",
     "—", "USD 500+/month enterprise",
     "Volume-based", "Account + zone token",
     "Most robust — residential IPs. For deep tier where reliability matters."),
    ("Brand/SEO", "Wikipedia MediaWiki API",
     "https://en.wikipedia.org/w/api.php",
     "REST/JSON", "Free", "—",
     "200 req/sec polite", "None",
     "Concept-page existence check. No key required."),
    ("Brand/SEO", "Wikidata SPARQL endpoint",
     "https://query.wikidata.org/sparql",
     "SPARQL/REST", "Free", "—",
     "60 req/min polite", "None",
     "Structured entity data. Useful for disambiguation."),
    ("Brand/SEO", "pytrends (unofficial Google Trends)",
     "https://github.com/GeneralMills/pytrends",
     "Python library", "Free / open source", "—",
     "Self-throttle ~1 req/sec", "None",
     "Unofficial. Use sparingly to avoid Google IP block. Fallback: SerpAPI Trends."),
    ("Brand/SEO", "Crunchbase Basic API",
     "https://data.crunchbase.com/",
     "REST/JSON", "—", "USD 49+/month",
     "Tier-dependent", "API key",
     "Funded-startup name conflicts. Paid only."),
    ("Brand/SEO", "Tracxn (no public API)",
     "https://tracxn.com/",
     "Manual UI only", "—", "Enterprise contract",
     "—", "—",
     "Manual lookup or enterprise data deal. Skip for v1."),

    # ──────────────── LINGUISTIC & NUMEROLOGY ────────────────
    ("CATEGORY", "LINGUISTIC, TRANSLITERATION & NUMEROLOGY", "", "", "", "", "", "", ""),
    ("Linguistic", "Bhashini API (Govt of India)",
     "https://bhashini.gov.in/ulca/",
     "REST/JSON", "Free (with registration)", "—",
     "Per-allocation", "Account token",
     "ULCA platform. Translation, transliteration, ASR for all 22 scheduled languages."),
    ("Linguistic", "Indic NLP Library",
     "https://github.com/anoopkunchukuttan/indic_nlp_library",
     "Python library", "Free / open source", "—",
     "Local", "None",
     "Self-host. Tokenization, normalization, transliteration."),
    ("Linguistic", "Aksharamukha (transliteration)",
     "http://aksharamukha.appspot.com/",
     "REST/JSON + library", "Free", "—",
     "Polite (~5/sec)", "None",
     "Best-in-class Indic script transliteration. ISO 15919 compliant."),
    ("Linguistic", "Google Cloud Translation API",
     "https://cloud.google.com/translate/docs",
     "REST/JSON", "500k chars/month free", "USD 20 per 1M chars",
     "Volume-based", "API key + GCP project",
     "Fallback for unusual languages."),
    ("Linguistic", "ICU4J / PyICU (transliteration)",
     "https://icu.unicode.org/",
     "Library", "Free / open source", "—",
     "Local", "None",
     "Industrial-strength Unicode transliteration. Use for normalisation."),
    ("Linguistic", "Custom Chaldean numerology engine",
     "internal://numerology",
     "Internal library (Python)", "Free (proprietary)", "—",
     "Local", "None",
     "Letter→digit mapping, compound, root, planet, industry-fit. Pure Python."),

    # ──────────────── INFRA: CAPTCHA, PROXIES, QUEUES ────────────────
    ("CATEGORY", "INFRASTRUCTURE — CAPTCHA, proxies, queues, real-time", "", "", "", "", "", "", ""),
    ("Infra", "2Captcha (CAPTCHA solving)",
     "https://2captcha.com/",
     "REST API", "—", "USD 2.99 per 1k reCAPTCHA v2 / USD 2.99 per 1k normal",
     "Volume", "API key",
     "Used by deep-scan tier. Average solve time 15–30s."),
    ("Infra", "Anti-Captcha",
     "https://anti-captcha.com/",
     "REST API", "—", "USD 2.00 per 1k normal / USD 2.00 per 1k reCAPTCHA v2",
     "Volume", "API key",
     "Cheaper alternative. Slightly higher failure rate."),
    ("Infra", "CapMonster Cloud",
     "https://capmonster.cloud/",
     "REST API", "—", "USD 0.50–1.50 per 1k",
     "Volume", "API key",
     "Cheapest. Use as failover if quality is acceptable."),
    ("Infra", "Bright Data residential proxies",
     "https://brightdata.com/",
     "HTTP/SOCKS5 proxy", "—", "USD 15.75 per GB pay-as-you-go",
     "Bandwidth-based", "Zone token",
     "Residential IPs for hardest sources. Use sparingly."),
    ("Infra", "Smartproxy",
     "https://smartproxy.com/",
     "HTTP/SOCKS5 proxy", "—", "USD 7 per GB starter",
     "Bandwidth-based", "Whitelisted IP + auth",
     "Cheaper residential. Good for MCA, IP India."),
    ("Infra", "Redis (cache + pub/sub + queue)",
     "https://redis.io/",
     "TCP", "Free / open source / Redis Cloud free tier",
     "USD 7+/month managed", "Self-hosted",
     "TLS recommended",
     "Self-host or use Upstash for serverless. Core dependency."),
    ("Infra", "BullMQ (job queue, Redis-backed)",
     "https://docs.bullmq.io/",
     "Node.js library", "Free / open source", "—",
     "Local", "Same as Redis",
     "Priority lanes, retries, rate limits, scheduling — all built in."),
    ("Infra", "Socket.IO (WebSocket server)",
     "https://socket.io/",
     "Node.js library", "Free / open source", "—",
     "Server resource", "Token or session",
     "Auto-reconnect, fallback to polling. Use Redis adapter for multi-node."),
    ("Infra", "Pusher Channels (managed real-time)",
     "https://pusher.com/channels/",
     "Hosted WebSocket service", "200k msgs/day free", "USD 49+/month",
     "Quota", "App key + secret",
     "Alternative to self-hosted Socket.IO if ops bandwidth limited."),
    ("Infra", "AWS API Gateway WebSocket",
     "https://aws.amazon.com/api-gateway/",
     "Managed WebSocket", "—", "USD 1.00 per million msgs",
     "Quota", "IAM",
     "Most cost-effective at scale. Plug into Lambda or EC2."),
    ("Infra", "AWS S3 (PDF report storage)",
     "https://aws.amazon.com/s3/",
     "Object storage", "Free 5GB / 12mo", "USD 0.023 per GB/month",
     "Quota", "IAM",
     "Pre-signed URLs for time-limited downloads."),
    ("Infra", "Cloudflare R2 (S3-compatible)",
     "https://www.cloudflare.com/products/r2/",
     "Object storage", "10GB free", "USD 0.015 per GB/month",
     "Quota", "API token",
     "Zero egress fees — cheaper than S3 for high download."),

    # ──────────────── AUTHENTICATION & BILLING ────────────────
    ("CATEGORY", "AUTHENTICATION & PAYMENTS", "", "", "", "", "", "", ""),
    ("Auth/Pay", "Firebase Authentication (phone + OTP)",
     "https://firebase.google.com/docs/auth/web/phone-auth",
     "SDK + REST", "Free up to 10k verifications/mo", "USD 0.01–0.06 per SMS",
     "Quota", "Firebase project config",
     "Phone OTP for sign-in. Works globally. Recaptcha-protected."),
    ("Auth/Pay", "MSG91 (SMS OTP — India)",
     "https://msg91.com/",
     "REST API", "100 free credits trial", "₹0.18–₹0.25 per SMS",
     "Volume", "Auth key",
     "Cheaper than Firebase for India-only SMS. DLT-registered."),
    ("Auth/Pay", "WhatsApp Business OTP (Meta)",
     "https://developers.facebook.com/docs/whatsapp/cloud-api",
     "REST API", "Free dev sandbox", "₹0.30–₹0.80 per OTP message (India)",
     "Quota", "WhatsApp Business account",
     "Cheaper than SMS in India. Preferred channel."),
    ("Auth/Pay", "Razorpay (payments)",
     "https://razorpay.com/docs/",
     "REST API + SDKs", "Test mode free", "2% transaction fee (UPI 0%)",
     "Volume", "API key + secret",
     "Best Indian payment gateway. UPI, cards, net banking, EMI."),
    ("Auth/Pay", "Stripe (international payments)",
     "https://stripe.com/docs",
     "REST API + SDKs", "Test mode free", "2.9% + ₹3 per transaction",
     "Volume", "API key + webhook secret",
     "For non-India payments and global agency clients."),
    ("Auth/Pay", "ClearTax / TaxAdda (GST e-invoice)",
     "https://cleartax.in/",
     "REST API", "—", "₹0.50–₹2.00 per invoice",
     "Volume", "API key",
     "Automate GST-compliant invoicing for ₹49 and ₹999 transactions."),
]

# Column widths
ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 32
ws.column_dimensions["C"].width = 38
ws.column_dimensions["D"].width = 22
ws.column_dimensions["E"].width = 22
ws.column_dimensions["F"].width = 26
ws.column_dimensions["G"].width = 16
ws.column_dimensions["H"].width = 20
ws.column_dimensions["I"].width = 50

# Write data
for i, row in enumerate(rows, start=5):
    is_cat = row[0] == "CATEGORY"
    for col, val in enumerate(row, 1):
        cell = ws.cell(row=i, column=col, value=val)
        if is_cat:
            style_category(cell)
            if col > 1:
                cell.value = ""
        else:
            # Colour the "Free Tier" / "Paid Tier" columns
            if col == 5 and val and val != "—":
                style_body(cell, fill=OK_BG)
                cell.font = Font(name="Arial", size=10, color=OK_INK, bold=True)
            elif col == 6 and val and val != "—":
                style_body(cell, fill=WARN_BG)
                cell.font = Font(name="Arial", size=10, color=WARN_INK)
            else:
                style_body(cell)
    # Merge category row
    if is_cat:
        ws.merge_cells(start_row=i, end_row=i, start_column=1, end_column=9)
        ws.row_dimensions[i].height = 22
    else:
        ws.row_dimensions[i].height = 46

# Freeze top header
ws.freeze_panes = "A5"


# ════════════════════════════════════════════════════════════════
# SHEET 2 — SCANNER MODULE OWNERSHIP
# ════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Scanner Modules")
ws2["A1"] = "Scanner Module Ownership Matrix"
ws2["A1"].font = Font(name="Arial", size=16, bold=True, color=INK)
ws2.merge_cells("A1:F1")
ws2.row_dimensions[1].height = 28

headers2 = ["Scanner Module", "Connectors Used", "Worker Pool Size", "Avg Response Time (ms)", "Cache TTL", "Owner / Squad"]
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=3, column=col, value=h)
    style_header(cell)
ws2.row_dimensions[3].height = 32

scanner_data = [
    ("Legal Scanner", "MCA, IP India, Copyright, GST, DPIIT, FSSAI, RBI, SEBI, IRDAI, Patent Office, Trademarkia (fallback)", 12, "1800-3500", "24h for register hits; 7d for misses", "Legal-Data squad"),
    ("Domain Probe", "WHOIS, RDAP, GoDaddy, Namecheap, WhoisXML, Domainr, INRegistry, Cloudflare", 8, "300-900", "5min for availability; 24h for pricing", "Infra squad"),
    ("Social Handle Bot", "Instagram, X, YouTube, LinkedIn, Facebook, Threads, Telegram, WhatsApp, Pinterest", 10, "400-1200", "30min for availability", "Social-Integrations squad"),
    ("Marketplace Crawler", "google-play-scraper, iTunes Search, Product Hunt, GitHub, Shopify DNS, ScrapingBee (Amazon, Flipkart)", 8, "600-2200", "12h for app listings; 24h for sellers", "Infra squad"),
    ("Brand & SEO Engine", "Google CSE, SerpAPI, ScrapingBee, Bright Data, Wikipedia, Wikidata, pytrends, Crunchbase", 6, "1200-3000", "6h for SERP; 24h for Wikipedia", "Brand-Intel squad"),
    ("Linguistic + Numerology", "Bhashini, Indic NLP, Aksharamukha, ICU4J, custom Chaldean engine", 4, "100-400", "Permanent (deterministic)", "NLP squad"),
]

# Column widths
ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 68
ws2.column_dimensions["C"].width = 18
ws2.column_dimensions["D"].width = 22
ws2.column_dimensions["E"].width = 30
ws2.column_dimensions["F"].width = 26

for i, row in enumerate(scanner_data, start=4):
    for col, val in enumerate(row, 1):
        cell = ws2.cell(row=i, column=col, value=val)
        style_body(cell)
    ws2.row_dimensions[i].height = 44

# Summary stats row
ws2.cell(row=4 + len(scanner_data) + 1, column=1, value="TOTAL WORKERS").font = Font(name="Arial", bold=True, size=10, color="FAF8F3")
ws2.cell(row=4 + len(scanner_data) + 1, column=1).fill = PatternFill("solid", start_color=INK)
ws2.cell(row=4 + len(scanner_data) + 1, column=1).border = border
ws2.cell(row=4 + len(scanner_data) + 1, column=3, value=f"=SUM(C4:C{3+len(scanner_data)})").font = Font(name="Arial", bold=True, size=11, color=ACCENT)
ws2.cell(row=4 + len(scanner_data) + 1, column=3).border = border
ws2.cell(row=4 + len(scanner_data) + 1, column=3).alignment = Alignment(horizontal="left", vertical="center")

ws2.freeze_panes = "A4"


# ════════════════════════════════════════════════════════════════
# SHEET 3 — COST MODEL (per scan)
# ════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Cost Model")
ws3["A1"] = "Per-Scan Cost Model — Free vs Deep Tier"
ws3["A1"].font = Font(name="Arial", size=16, bold=True, color=INK)
ws3.merge_cells("A1:E1")
ws3.row_dimensions[1].height = 28

ws3["A2"] = "Adjust the Blue input cells to recompute. All other cells are formulas."
ws3["A2"].font = Font(name="Arial", size=10, italic=True, color=INK3)
ws3.merge_cells("A2:E2")

headers3 = ["Line item", "Unit cost (₹)", "Usage per Free scan", "Usage per Deep scan", "Cost per Deep scan (₹)"]
for col, h in enumerate(headers3, 1):
    cell = ws3.cell(row=4, column=col, value=h)
    style_header(cell)
ws3.row_dimensions[4].height = 32

cost_rows = [
    ("CAPTCHA solve (2Captcha)", 0.25, 0, 3, "=B5*D5"),
    ("Residential proxy bandwidth (per 100KB)", 0.13, 0, 8, "=B6*D6"),
    ("SerpAPI call", 1.0, 0, 1, "=B7*D7"),
    ("Google CSE call", 0.40, 0, 1, "=B8*D8"),
    ("WhoisXML call", 0.04, 1, 1, "=B9*D9"),
    ("SMS OTP (MSG91)", 0.22, 0, 0, "=B10*D10"),
    ("Compute time per scan (server-seconds)", 0.05, 4, 12, "=B11*D11"),
    ("PDF generation + storage", 0.20, 0, 1, "=B12*D12"),
    ("Bandwidth out (Cloudflare R2 — free)", 0.0, 0, 1, "=B13*D13"),
]
for i, row in enumerate(cost_rows, start=5):
    for col, val in enumerate(row, 1):
        cell = ws3.cell(row=i, column=col, value=val)
        if col == 2:
            cell.font = Font(name="Arial", size=10, color="0000FF", bold=True)
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = "₹#,##0.00"
        elif col == 5:
            cell.font = Font(name="Arial", size=10, color=INK)
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = "₹#,##0.00"
        else:
            style_body(cell)
        cell.border = border
    ws3.row_dimensions[i].height = 22

# Totals
total_row = 5 + len(cost_rows)
ws3.cell(row=total_row, column=1, value="TOTAL COGS per Deep scan").font = Font(name="Arial", size=11, bold=True, color="FAF8F3")
ws3.cell(row=total_row, column=1).fill = PatternFill("solid", start_color=ACCENT)
ws3.cell(row=total_row, column=1).border = border
ws3.merge_cells(start_row=total_row, end_row=total_row, start_column=1, end_column=4)
ws3.cell(row=total_row, column=5, value=f"=SUM(E5:E{total_row-1})")
ws3.cell(row=total_row, column=5).font = Font(name="Arial", size=11, bold=True, color="FAF8F3")
ws3.cell(row=total_row, column=5).fill = PatternFill("solid", start_color=ACCENT)
ws3.cell(row=total_row, column=5).number_format = "₹#,##0.00"
ws3.cell(row=total_row, column=5).border = border
ws3.cell(row=total_row, column=5).alignment = Alignment(horizontal="right")

# Margin
ws3.cell(row=total_row + 2, column=1, value="Price (Deep Scan)").font = Font(name="Arial", size=11, bold=True)
ws3.cell(row=total_row + 2, column=5, value=49.0).font = Font(name="Arial", size=11, color="0000FF", bold=True)
ws3.cell(row=total_row + 2, column=5).number_format = "₹#,##0.00"
ws3.cell(row=total_row + 2, column=5).alignment = Alignment(horizontal="right")

ws3.cell(row=total_row + 3, column=1, value="Gross profit per Deep Scan (₹)").font = Font(name="Arial", size=11, bold=True)
ws3.cell(row=total_row + 3, column=5, value=f"=E{total_row+2}-E{total_row}").font = Font(name="Arial", size=11, bold=True, color=OK_INK)
ws3.cell(row=total_row + 3, column=5).number_format = "₹#,##0.00"
ws3.cell(row=total_row + 3, column=5).alignment = Alignment(horizontal="right")

ws3.cell(row=total_row + 4, column=1, value="Gross margin %").font = Font(name="Arial", size=11, bold=True)
ws3.cell(row=total_row + 4, column=5, value=f"=(E{total_row+2}-E{total_row})/E{total_row+2}").font = Font(name="Arial", size=11, bold=True, color=OK_INK)
ws3.cell(row=total_row + 4, column=5).number_format = "0.0%"
ws3.cell(row=total_row + 4, column=5).alignment = Alignment(horizontal="right")

ws3.column_dimensions["A"].width = 44
ws3.column_dimensions["B"].width = 22
ws3.column_dimensions["C"].width = 24
ws3.column_dimensions["D"].width = 24
ws3.column_dimensions["E"].width = 24

# Save
wb.save(OUT)
print("Wrote:", OUT)
